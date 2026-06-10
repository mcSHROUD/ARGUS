from __future__ import annotations

import io
import json
import re
from datetime import date as date_type
from pathlib import Path

import cv2
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from argus.classifier import DefectClassifier
from argus.config import load_config
from argus.detector import DefectDetector
from argus.storage import Storage

config = load_config()
storage = Storage(config.storage.db_path, config.storage.images_dir)

templates = Jinja2Templates(directory=str(Path(__file__).parent / "templates"))
app = FastAPI(title="ARGUS review")

_detector: DefectDetector | None = None
_classifier: DefectClassifier | None = None
_inspect_history: list[dict] = []


def get_detector() -> DefectDetector | None:
    global _detector
    if _detector is None:
        try:
            d = DefectDetector(config.detector)
            d.load()
            _detector = d
        except Exception:
            return None
    return _detector


def get_classifier() -> DefectClassifier | None:
    global _classifier
    if _classifier is None:
        try:
            clf_path = Path(config.detector.model_path).parent / "classifier.pt"
            c = DefectClassifier(clf_path)
            c.load()
            _classifier = c
        except Exception:
            return None
    return _classifier


@app.get("/", response_class=HTMLResponse)
def index(request: Request, date: str | None = None) -> HTMLResponse:
    if date is None:
        date = date_type.today().isoformat()
    stats = storage.get_daily_stats(date)
    hourly = storage.get_hourly_counts(date)
    hourly_by_hour = {r["hour"]: r for r in hourly}
    hourly_full = [
        hourly_by_hour.get(h, {"hour": h, "defects": 0, "good": 0})
        for h in range(24)
    ]
    defect_cups = storage.list_cups(limit=50, verdict="defect", date=date)
    return templates.TemplateResponse(
        request,
        "index.html",
        {
            "stats": stats,
            "hourly_json": json.dumps(hourly_full),
            "cups": defect_cups,
            "date": date,
        },
    )


@app.get("/cup/{cup_id}", response_class=HTMLResponse)
def cup_detail(request: Request, cup_id: int) -> HTMLResponse:
    cup = storage.get_cup(cup_id)
    if cup is None:
        raise HTTPException(404, "cup not found")
    return templates.TemplateResponse(request, "cup.html", {"cup": cup})


@app.get("/image")
def image(path: str) -> FileResponse:
    p = Path(path)
    images_root = Path(config.storage.images_dir).resolve()
    try:
        p.resolve().relative_to(images_root)
    except ValueError as e:
        raise HTTPException(403, "path outside images_dir") from e
    if not p.exists():
        raise HTTPException(404, "image not found")
    return FileResponse(p)


@app.get("/export")
def export_excel(date: str | None = None) -> StreamingResponse:
    if date is None:
        date = date_type.today().isoformat()
    cups = storage.list_cups(limit=10000, date=date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"ARGUS {date}"

    headers = ["ID", "Время", "Вердикт", "Score", "Камера 1", "Примечание"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for cup in cups:
        ws.append([
            cup.id,
            cup.ts.strftime("%Y-%m-%d %H:%M:%S"),
            cup.verdict,
            round(cup.score, 3),
            cup.cam1_path or "",
            cup.notes or "",
        ])
        if cup.verdict == "defect":
            for cell in ws[ws.max_row]:
                cell.fill = red_fill

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"argus_{date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/inspect", response_class=HTMLResponse)
def inspect_form(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(request, "inspect.html", {"results": None})


@app.post("/inspect", response_class=HTMLResponse)
async def inspect_photos(request: Request, files: list[UploadFile] = File(...)) -> HTMLResponse:
    clf = get_classifier()
    det = get_detector()
    if clf is None and det is None:
        raise HTTPException(503, "Ни классификатор, ни PatchCore не загружены.")

    use_classifier = clf is not None
    threshold_display = round(clf.threshold, 2) if use_classifier else round(det.threshold, 2)  # type: ignore[union-attr]
    mode = "EfficientNet-B0" if use_classifier else "PatchCore"

    allowed = {".jpg", ".jpeg", ".png"}
    seen: set[str] = set()
    results = []
    for upload in files:
        name = Path(upload.filename or "").name
        suffix = Path(upload.filename or "").suffix.lower()
        if suffix not in allowed or name in seen:
            continue
        seen.add(name)
        data = await upload.read()
        arr = np.frombuffer(data, np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if img is None:
            continue

        if use_classifier:
            prob, is_defect, anomaly_map = clf.predict_with_gradcam(img)  # type: ignore[union-attr]
            score = round(prob * 100, 2)         # 0-100 scale for display
        else:
            result = det.infer([img], [])        # type: ignore[union-attr]
            score = round(result.score, 2)
            is_defect = result.is_defect
            anomaly_map = result.anomaly_map_cam1

        thumb, full, dtype = _prepare_images(img, anomaly_map, is_defect)
        results.append({
            "filename": name,
            "score": score,
            "is_defect": is_defect,
            "defect_type": dtype,
            "img_b64": thumb,
            "img_full_b64": full,
        })

    results.sort(key=lambda x: x["score"], reverse=True)

    if results:
        from datetime import datetime
        _inspect_history.insert(0, {
            "ts": datetime.now().strftime("%H:%M:%S"),
            "results": results,
            "threshold": threshold_display,
            "mode": mode,
        })

    return templates.TemplateResponse(
        request, "inspect.html",
        {
            "results": results,
            "threshold": threshold_display,
            "mode": mode,
            "history": _inspect_history[1:],
        },
    )


def _prepare_images(
    img: np.ndarray,
    anomaly_map: np.ndarray | None,
    is_defect: bool,
) -> tuple[str, str, str]:
    """Return (thumb_b64, full_b64, defect_type)."""
    import base64

    display = img
    defect_type = ""

    if is_defect and anomaly_map is not None:
        am = anomaly_map.squeeze()
        h, w = img.shape[:2]
        am_r = cv2.resize(am, (w, h))

        a_min, a_max = float(am_r.min()), float(am_r.max())
        if a_max > a_min:
            am_norm = ((am_r - a_min) / (a_max - a_min) * 255).astype(np.uint8)
        else:
            am_norm = np.zeros((h, w), dtype=np.uint8)

        heatmap = cv2.applyColorMap(am_norm, cv2.COLORMAP_JET)
        display = cv2.addWeighted(img, 0.50, heatmap, 0.50, 0)

        thresh_val = int(np.percentile(am_norm, 80))
        _, binary = cv2.threshold(am_norm, thresh_val, 255, cv2.THRESH_BINARY)
        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if contours:
            cv2.drawContours(display, contours, -1, (0, 255, 0), 2)
            defect_type = _classify_defect(binary, h, w)

    elif is_defect and anomaly_map is None:
        defect_type = "дефект"

    def _enc(src: np.ndarray, max_side: int) -> str:
        fh, fw = src.shape[:2]
        if max(fh, fw) > max_side:
            s = max_side / max(fh, fw)
            src = cv2.resize(src, (int(fw * s), int(fh * s)))
        _, buf = cv2.imencode(".jpg", src, [cv2.IMWRITE_JPEG_QUALITY, 85])
        return base64.b64encode(buf).decode()

    return _enc(display, 120), _enc(display, 900), defect_type


def _classify_defect_image(img: np.ndarray) -> str:
    """Classify defect type from raw image using edge heuristics (no anomaly map)."""
    h, w = img.shape[:2]
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(gray)
    edges = cv2.Canny(enhanced, 40, 120)

    # Seam: strong edge concentration near top/bottom border
    border_h = int(h * 0.14)
    border_edges = int(edges[:border_h, :].sum() + edges[h - border_h:, :].sum())
    total_edges = int(edges.sum()) or 1
    if border_edges / total_edges > 0.45:
        return "дефект шва"

    # Scratch: long straight lines via Hough
    min_len = int(min(h, w) * 0.12)
    lines = cv2.HoughLinesP(edges, 1, np.pi / 180,
                             threshold=25, minLineLength=min_len,
                             maxLineGap=int(min_len * 0.4))
    if lines is not None:
        for seg in lines:
            x1, y1, x2, y2 = seg[0]
            length = float(np.hypot(x2 - x1, y2 - y1))
            if length > min(h, w) * 0.15:
                return "царапина"

    return "вмятина"


def _classify_defect(mask: np.ndarray, img_h: int, img_w: int) -> str:
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return "неопределён"
    c = max(contours, key=cv2.contourArea)
    area = cv2.contourArea(c)
    if area < 80:
        return "неопределён"

    x, y, bw, bh = cv2.boundingRect(c)
    if y < img_h * 0.13 or (y + bh) > img_h * 0.87:
        return "дефект шва"

    sides = sorted([bw, bh])
    aspect = sides[1] / max(sides[0], 1)
    perimeter = cv2.arcLength(c, True)
    circularity = 4 * 3.14159 * area / max(perimeter ** 2, 1e-6)

    if aspect >= 2.5:
        return "царапина"
    if circularity >= 0.45:
        return "вмятина"
    if aspect >= 1.5:
        return "царапина"
    return "вмятина"


@app.get("/cleanup", response_class=HTMLResponse)
def cleanup_page(request: Request) -> HTMLResponse:
    days = storage.get_available_days()
    return templates.TemplateResponse(request, "cleanup.html", {"days": days})


@app.delete("/cleanup/{date}")
def cleanup_day(date: str) -> dict:
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", date):
        raise HTTPException(400, "invalid date format")
    result = storage.delete_day(date)
    return result


@app.post("/inspect/clear")
def clear_history() -> dict:
    _inspect_history.clear()
    return {"ok": True}


@app.get("/health")
def health() -> dict:
    return {"status": "ok"}
